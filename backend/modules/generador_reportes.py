import os
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from backend.models.database import Pedido, PagoPlataforma, Periodo

VERDE = "FF00C853"
ROJO = "FFFF1744"
AMARILLO = "FFFFAB00"
AZUL = "FF1565C0"
GRIS = "FFF5F5F5"
BLANCO = "FFFFFFFF"

def estilo_encabezado(cell, color=AZUL):
    cell.font = Font(bold=True, color=BLANCO, size=11)
    cell.fill = PatternFill(fill_type="solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

def estilo_celda(cell, bold=False, color=None, num_format=None):
    cell.font = Font(bold=bold, size=10)
    if color:
        cell.fill = PatternFill(fill_type="solid", fgColor=color)
    cell.alignment = Alignment(vertical="center")
    if num_format:
        cell.number_format = num_format
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

def ajustar_columnas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

def agregar_titulo(ws, titulo, subtitulo, periodo):
    ws.merge_cells("A1:H1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14, color=BLANCO)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=AZUL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Periodo: {periodo.fecha_inicio.strftime(chr(37)+chr(100)+chr(47)+chr(37)+chr(109)+chr(47)+chr(37)+chr(89))} al {periodo.fecha_corte.strftime(chr(37)+chr(100)+chr(47)+chr(37)+chr(109)+chr(47)+chr(37)+chr(89))} | {subtitulo}"
    ws["A2"].font = Font(italic=True, size=10, color="FF555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

def reporte_a_pagados(ws, pedidos, periodo):
    agregar_titulo(ws, "A. PEDIDOS PAGADOS EN EL PERIODO", "Todos los pagos recibidos", periodo)
    headers = ["# Pedido","Cliente","Plataforma","Fecha Pedido","Fecha Pago","Valor Total","Comision","Estado"]
    for i, h in enumerate(headers, 1):
        estilo_encabezado(ws.cell(3, i, h))
    ws.row_dimensions[3].height = 20
    for r, p in enumerate(pedidos, 4):
        color = GRIS if r % 2 == 0 else BLANCO
        ws.cell(r, 1, p.numero_pedido)
        ws.cell(r, 2, p.cliente_nombre or "")
        ws.cell(r, 3, p.plataforma_pago or "")
        ws.cell(r, 4, p.fecha_pedido.strftime("%d/%m/%Y") if p.fecha_pedido else "")
        ws.cell(r, 5, p.fecha_pago_real.strftime("%d/%m/%Y") if p.fecha_pago_real else "")
        ws.cell(r, 6, p.valor_total)
        ws.cell(r, 7, p.comision_aplicada or 0)
        ws.cell(r, 8, "Pagado")
        for c in range(1, 9):
            estilo_celda(ws.cell(r, c), color=color)
        ws.cell(r, 6).number_format = "0,##0"
        ws.cell(r, 7).number_format = "0,##0"
        ws.cell(r, 8).font = Font(bold=True, color="FF" + VERDE[2:])

def reporte_b_mora(ws, pedidos, periodo):
    agregar_titulo(ws, "B. PEDIDOS PAGADOS CON MORA", "Pagos recibidos despues del plazo", periodo)
    headers = ["# Pedido","Cliente","Plataforma","Fecha Esperada","Fecha Real","Dias Mora","Valor Total","Estado"]
    for i, h in enumerate(headers, 1):
        estilo_encabezado(ws.cell(3, i, h), color="FFFF6F00")
    for r, p in enumerate(pedidos, 4):
        color = "FFFFF8E1" if r % 2 == 0 else BLANCO
        ws.cell(r, 1, p.numero_pedido)
        ws.cell(r, 2, p.cliente_nombre or "")
        ws.cell(r, 3, p.plataforma_pago or "")
        ws.cell(r, 4, p.fecha_pago_esperada.strftime("%d/%m/%Y") if p.fecha_pago_esperada else "")
        ws.cell(r, 5, p.fecha_pago_real.strftime("%d/%m/%Y") if p.fecha_pago_real else "")
        ws.cell(r, 6, p.dias_mora or 0)
        ws.cell(r, 7, p.valor_total)
        ws.cell(r, 8, f"{p.dias_mora} dias de mora")
        for c in range(1, 9):
            estilo_celda(ws.cell(r, c), color=color)
        ws.cell(r, 6).font = Font(bold=True, color="FFFF1744")
        ws.cell(r, 7).number_format = "0,##0"

def reporte_c_cxc(ws, pedidos, periodo):
    agregar_titulo(ws, "C. CUENTAS POR COBRAR AL CIERRE", "Pedidos aun sin pago al final del periodo", periodo)
    headers = ["# Pedido","Cliente","Plataforma","Fecha Pedido","Fecha Esperada","Valor","Dias Pendiente","Prioridad"]
    for i, h in enumerate(headers, 1):
        estilo_encabezado(ws.cell(3, i, h), color="FF880E4F")
    today = datetime.now().date()
    for r, p in enumerate(pedidos, 4):
        dias = (today - p.fecha_pedido.date()).days if p.fecha_pedido else 0
        prioridad = "URGENTE" if dias > 30 else "ALTA" if dias > 15 else "NORMAL"
        color_prior = "FFFFEBEE" if prioridad == "URGENTE" else "FFFFF8E1" if prioridad == "ALTA" else BLANCO
        ws.cell(r, 1, p.numero_pedido)
        ws.cell(r, 2, p.cliente_nombre or "")
        ws.cell(r, 3, p.plataforma_pago or "")
        ws.cell(r, 4, p.fecha_pedido.strftime("%d/%m/%Y") if p.fecha_pedido else "")
        ws.cell(r, 5, p.fecha_pago_esperada.strftime("%d/%m/%Y") if p.fecha_pago_esperada else "")
        ws.cell(r, 6, p.valor_total)
        ws.cell(r, 7, dias)
        ws.cell(r, 8, prioridad)
        for c in range(1, 9):
            estilo_celda(ws.cell(r, c), color=color_prior)
        ws.cell(r, 6).number_format = "0,##0"

def reporte_d_comisiones(ws, pagos, periodo):
    agregar_titulo(ws, "D. COMISIONES POR MEDIO DE PAGO", "Valor cobrado vs valor sin comision", periodo)
    headers = ["Plataforma","Medio de Pago","# Pagos","Valor Bruto","Valor Neto","Comision","Tasa %","Fecha"]
    for i, h in enumerate(headers, 1):
        estilo_encabezado(ws.cell(3, i, h), color="FF1B5E20")
    from backend.modules.credenciales import calcular_valor_bruto
    for r, p in enumerate(pagos, 4):
        color = "FFF1F8E9" if r % 2 == 0 else BLANCO
        try:
            info = calcular_valor_bruto(p.plataforma, p.valor_neto, p.metodo_pago or "default")
            bruto = info["valor_bruto"]
            comision = info["comision_total"]
            tasa = info["tasa_total"]
        except Exception:
            bruto = p.valor_neto
            comision = 0
            tasa = 0
        ws.cell(r, 1, p.plataforma)
        ws.cell(r, 2, p.metodo_pago or "default")
        ws.cell(r, 3, 1)
        ws.cell(r, 4, bruto)
        ws.cell(r, 5, p.valor_neto)
        ws.cell(r, 6, comision)
        ws.cell(r, 7, tasa)
        ws.cell(r, 8, p.fecha_pago.strftime("%d/%m/%Y") if p.fecha_pago else "")
        for c in range(1, 9):
            estilo_celda(ws.cell(r, c), color=color)
        for c in [4, 5, 6]:
            ws.cell(r, c).number_format = "0,##0"
        ws.cell(r, 7).number_format = "0.00%"

def generar_reportes_excel(db: Session, periodo_id: int, carpeta_salida: str) -> str:
    periodo = db.query(Periodo).filter_by(id=periodo_id).first()
    if not periodo:
        raise ValueError("Periodo no encontrado")
    pedidos_pagados = db.query(Pedido).filter(
        Pedido.periodo_id == periodo_id,
        Pedido.estado_conciliacion == "pagado"
    ).all()
    pedidos_mora = [p for p in pedidos_pagados if p.dias_mora and p.dias_mora > 0]
    pedidos_cxc = db.query(Pedido).filter(
        Pedido.periodo_id == periodo_id,
        Pedido.estado_conciliacion == "pendiente"
    ).all()
    pagos = db.query(PagoPlataforma).filter(
        PagoPlataforma.periodo_id == periodo_id
    ).all()
    wb = Workbook()
    wb.remove(wb.active)
    ws_a = wb.create_sheet("A. Pagados")
    reporte_a_pagados(ws_a, pedidos_pagados, periodo)
    ajustar_columnas(ws_a)
    ws_b = wb.create_sheet("B. Mora")
    reporte_b_mora(ws_b, pedidos_mora, periodo)
    ajustar_columnas(ws_b)
    ws_c = wb.create_sheet("C. CxC Cierre")
    reporte_c_cxc(ws_c, pedidos_cxc, periodo)
    ajustar_columnas(ws_c)
    ws_d = wb.create_sheet("D. Comisiones")
    reporte_d_comisiones(ws_d, pagos, periodo)
    ajustar_columnas(ws_d)
    os.makedirs(carpeta_salida, exist_ok=True)
    fecha_hoy = datetime.now().strftime("%Y%m%d_%H%M")
    ruta = os.path.join(carpeta_salida, f"conciliacion_{periodo_id}_{fecha_hoy}.xlsx")
    wb.save(ruta)
    return ruta