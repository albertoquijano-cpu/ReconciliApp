import os, re
from datetime import datetime
from sqlalchemy.orm import Session
from backend.models.database import PagoPlataforma

ORIGEN_KEYWORDS = {
    "shopify": ["shopify", "stripe"],
    "bold": ["bold"],
    "wompi": ["wompi", "adelante soluci"],
    "mercadopago": ["mercado pago", "mercadopago"],
    "paypal": ["paypal"],
    "addi": ["addi"],
    "sistecredito": ["sistecredito"],
    "nequi": ["nequi"],
    "qr": ["qr", "codigo qr", "pago qr"],
    "pse": ["pse", "ach"],
    "tarjeta": ["credibanco", "redeban", "visa", "mastercard"],
    "transferencia": ["transferencia", "trf"],
    "consignacion": ["consignacion"],
}

def identificar_origen(desc, tipo=None):
    if tipo:
        t = tipo.lower()
        if t == "addi": return "addi"
        if t == "wompi": return "wompi"
        if t == "nequi": return "nequi"
        if t == "qr": return "qr"
    if not desc: return "bancolombia"
    d = desc.lower()
    for origen, kws in ORIGEN_KEYWORDS.items():
        for kw in kws:
            if kw in d: return origen
    return "bancolombia"

def limpiar_valor(txt):
    if not txt: return 0.0
    l = re.sub(r"[^\d,.]", "", str(txt).strip())
    if "," in l and "." in l: l = l.replace(".","").replace(",",".")
    elif "," in l: l = l.replace(",",".")
    try: return float(l)
    except: return 0.0

def parsear_fecha(txt):
    if not txt: return None
    if isinstance(txt, datetime): return txt
    for fmt in ["%d/%m/%Y","%Y-%m-%d","%d-%m-%Y","%Y/%m/%d"]:
        try: return datetime.strptime(str(txt).strip()[:10], fmt)
        except: continue
    return None

def parsear_extracto_bancolombia(ruta):
    ext = os.path.splitext(ruta)[1].lower()
    if ext in [".xlsx", ".xls"]:
        return _parsear_excel(ruta)
    else:
        return _parsear_csv(ruta)

def _parsear_excel(ruta):
    import openpyxl
    wb = openpyxl.load_workbook(ruta)
    # Buscar hoja con movimientos
    hoja = None
    for nombre in wb.sheetnames:
        if any(k in nombre.lower() for k in ["mov", "transac", "extracto", "movimiento"]):
            hoja = wb[nombre]
            break
    if not hoja:
        hoja = wb.active

    # Leer encabezados
    headers = [str(c.value).strip() if c.value else "" for c in list(hoja.rows)[0]]

    # Encontrar columnas
    def col_idx(opciones):
        for i, h in enumerate(headers):
            for op in opciones:
                if op.lower() == h.lower(): return i
        return None

    idx_fecha    = col_idx(["Fecha", "FECHA", "Fecha Transaccion"])
    idx_valor    = col_idx(["Valor", "VALOR", "Credito", "Abono", "Deposito"])
    idx_concepto = col_idx(["Concepto", "CONCEPTO", "Descripcion", "DESCRIPCION"])
    idx_tipo     = col_idx(["Tipo", "TIPO", "Origen"])

    movs = []
    for row in hoja.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        valor = row[idx_valor] if idx_valor is not None else 0
        if not valor or float(valor) <= 0: continue
        fecha = row[idx_fecha] if idx_fecha is not None else None
        concepto = row[idx_concepto] if idx_concepto is not None else ""
        tipo = row[idx_tipo] if idx_tipo is not None else ""
        movs.append({
            "fecha": parsear_fecha(fecha),
            "descripcion": str(concepto) if concepto else "",
            "referencia": "",
            "valor": float(valor),
            "origen": identificar_origen(str(concepto) if concepto else "", str(tipo) if tipo else ""),
        })
    return movs

def _parsear_csv(ruta):
    import csv
    from io import StringIO
    contenido = None
    for enc in ["utf-8-sig","latin-1","cp1252"]:
        try:
            with open(ruta, encoding=enc) as f: lines = f.readlines()
            contenido = lines
            break
        except: continue
    if not contenido: return []
    header_row = 0
    for i, line in enumerate(contenido):
        if any(c.lower() in line.lower() for c in ["fecha","debito","credito"]):
            header_row = i
            break
    reader = csv.DictReader(StringIO("".join(contenido[header_row:])))
    headers = reader.fieldnames or []
    def encontrar_col(opciones):
        for h in headers:
            for op in opciones:
                if op.lower() == h.lower().strip(): return h
        return None
    col_f = encontrar_col(["Fecha", "FECHA", "Fecha Transaccion"])
    col_d = encontrar_col(["Descripcion", "DESCRIPCION", "Concepto"])
    col_r = encontrar_col(["Referencia", "REFERENCIA", "No. Comprobante"])
    col_c = encontrar_col(["Credito", "CREDITO", "Deposito", "Abono", "Valor"])
    movs = []
    for row in reader:
        val = limpiar_valor(row.get(col_c, 0)) if col_c else 0
        if val <= 0: continue
        desc = row.get(col_d, "") if col_d else ""
        movs.append({
            "fecha": parsear_fecha(row.get(col_f,"")) if col_f else None,
            "descripcion": desc,
            "referencia": row.get(col_r,"") if col_r else "",
            "valor": val,
            "origen": identificar_origen(desc),
        })
    return movs

def guardar_extracto_en_bd(db, periodo_id, movimientos):
    guardados = 0
    for m in movimientos:
        db.add(PagoPlataforma(
            periodo_id=periodo_id,
            plataforma=m["origen"],
            fecha_pago=m["fecha"],
            valor_neto=m["valor"],
            valor_bruto=m["valor"],
            referencia=m["referencia"],
            descripcion=m["descripcion"],
            metodo_pago=m["origen"],
            estado_asignacion="sin_asignar",
        ))
        guardados += 1
    db.commit()
    return guardados
